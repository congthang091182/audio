import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# Hàm để ghi dữ liệu vào file Excel mẫu
def write_to_excel_template9(template_path, df1, df2,df3,df4,df5,df6, startrow, startcol, output_path):
    """
    Ghi dữ liệu vào file Excel mẫu.

    Parameters:
        template_path (str): Đường dẫn đến file Excel mẫu.
        df1 (pd.DataFrame): DataFrame thứ nhất.
        df2 (pd.DataFrame): DataFrame thứ hai.
        startrow (int): Dòng bắt đầu ghi dữ liệu.
        startcol (int): Cột bắt đầu ghi dữ liệu.
        output_path (str): Đường dẫn để lưu file Excel kết quả.
    """
     # Đọc file Excel mẫu
    book = load_workbook(template_path)
    
    # Ghi dữ liệu vào sheet đầu tiên (Sheet1)
    sheet1 = book[book.sheetnames[0]]  # Lấy sheet đầu tiên
    for r_idx, row in enumerate(df1.values, start=startrow):
        for c_idx, value in enumerate(row, start=startcol):
            sheet1.cell(row=r_idx, column=c_idx, value=value)
    
    # Ghi dữ liệu vào sheet thứ hai (Sheet2)
    if len(book.sheetnames) > 1:
        sheet2 = book[book.sheetnames[1]]  # Lấy sheet thứ hai
        for r_idx, row in enumerate(df2.values, start=startrow):
            for c_idx, value in enumerate(row, start=startcol):
                sheet2.cell(row=r_idx, column=c_idx, value=value)
    # Ghi dữ liệu vào sheet thứ hai (Sheet3)
    if len(book.sheetnames) > 1:
        sheet3 = book[book.sheetnames[2]]  # Lấy sheet thứ hai
        for r_idx, row in enumerate(df3.values, start=startrow):
            for c_idx, value in enumerate(row, start=startcol):
                sheet3.cell(row=r_idx, column=c_idx, value=value) 
     # Ghi dữ liệu vào sheet thứ hai (Sheet)
    if len(book.sheetnames) > 1:
        sheet4 = book[book.sheetnames[3]]  # Lấy sheet thứ hai
        for r_idx, row in enumerate(df4.values, start=startrow):
            for c_idx, value in enumerate(row, start=startcol):
                sheet4.cell(row=r_idx, column=c_idx, value=value)     
    if len(book.sheetnames) > 1:
        sheet5 = book[book.sheetnames[4]]  # Lấy sheet thứ hai
        for r_idx, row in enumerate(df5.values, start=startrow):
            for c_idx, value in enumerate(row, start=startcol):
                sheet5.cell(row=r_idx, column=c_idx, value=value)  
    if len(book.sheetnames) > 1:
        sheet6 = book[book.sheetnames[5]]  # Lấy sheet thứ hai
        for r_idx, row in enumerate(df6.values, start=startrow):
            for c_idx, value in enumerate(row, start=startcol):
                sheet6.cell(row=r_idx, column=c_idx, value=value)          
    else:
        st.warning("File Excel mẫu chỉ có một sheet. Dữ liệu df2 sẽ không được ghi.")

    # Lưu file Excel
    book.save(output_path)

# # Tạo giao diện Streamlit
# st.title("Xuất dữ liệu vào file Excel mẫu")

# # Tạo DataFrame mẫu
# df1 = pd.DataFrame({
#     'Column1': [1, 2, 3],
#     'Column2': ['A', 'B', 'C']
# })

# df2 = pd.DataFrame({
#     'ColumnA': [4, 5, 6],
#     'ColumnB': ['X', 'Y', 'Z']
# })

# # Hiển thị DataFrame trong Streamlit
# st.write("DataFrame 1:")
# st.write(df1)

# st.write("DataFrame 2:")
# st.write(df2)

# # Tải lên file Excel mẫu
# uploaded_file = st.file_uploader("Tải lên file Excel mẫu", type=["xlsx"])

# if uploaded_file is not None:
#     # Nhập vị trí bắt đầu ghi dữ liệu
#     st.write("Nhập vị trí bắt đầu ghi dữ liệu:")
#     startrow = st.number_input("Dòng bắt đầu", min_value=0, value=15)
#     startcol = st.number_input("Cột bắt đầu", min_value=0, value=2)

#     # Tạo nút để xuất dữ liệu vào file Excel
#     if st.button("Xuất dữ liệu vào Excel"):
#         output_path = "output.xlsx"
#         write_to_excel_template(uploaded_file, df1, df2, startrow, startcol, output_path)
        
#         # Hiển thị liên kết tải xuống
#         with open(output_path, "rb") as file:
#             btn = st.download_button(
#                 label="Tải xuống file Excel đã cập nhật",
#                 data=file,
#                 file_name=output_path,
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )