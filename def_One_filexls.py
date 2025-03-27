import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# Hàm để ghi dữ liệu vào file Excel mẫu
def write_to_excel_template_one(template_path, df1, startrow, startcol, output_path):
    """
    Ghi dữ liệu vào file Excel mẫu.

    Parameters:
        template_path (str): Đường dẫn đến file Excel mẫu.
        df1 (pd.DataFrame): DataFrame chứa dữ liệu.
        startrow (int): Dòng bắt đầu ghi dữ liệu.
        startcol (int): Cột bắt đầu ghi dữ liệu.
        output_path (str): Đường dẫn để lưu file Excel kết quả.
    """
    if df1.empty:
        st.warning("DataFrame rỗng, không có dữ liệu để xuất Excel.")
        return

    # Đọc file Excel mẫu
    book = load_workbook(template_path)
    
    # Ghi dữ liệu vào sheet đầu tiên (Sheet1)
    sheet1 = book[book.sheetnames[0]]  # Lấy sheet đầu tiên
    
    # Chuẩn hóa dữ liệu trước khi ghi
    df1_cleaned = df1.copy()
    for col in df1_cleaned.columns:
        # Sử dụng pandas.api.types.is_string_dtype để kiểm tra kiểu chuỗi
        if pd.api.types.is_string_dtype(df1_cleaned[col]):
            df1_cleaned[col] = df1_cleaned[col].astype(str).str.strip()
        else:
            if not df1_cleaned[col].empty and df1_cleaned[col].notna().any():
                df1_cleaned[col] = pd.to_numeric(df1_cleaned[col], errors='coerce')

    # Ghi dữ liệu vào sheet
    for r_idx, row in enumerate(df1_cleaned.values, start=startrow):
        for c_idx, value in enumerate(row, start=startcol):
            sheet1.cell(row=r_idx, column=c_idx, value=value)

    # Lưu file Excel
    book.save(output_path)