import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

def export_pos_to_DVUT_excel(df, template_file_path, output_file_path="output_pos_data_fixed.xlsx"):
    """
    Xuất dữ liệu POS từ DataFrame ra file Excel với các sheet riêng cho từng POS.
    
    Parameters:
    - df (pd.DataFrame): DataFrame chứa dữ liệu với các cột 'POS', 'DVUT', 'TENDV'.
    - template_file_path (str): Đường dẫn đến file mẫu Excel.
    - output_file_path (str): Đường dẫn đến file Excel kết quả.
    
    Returns:
    - str: Đường dẫn đến file kết quả nếu thành công, None nếu thất bại.
    """
    if df.empty:
        st.warning("DataFrame rỗng, không có dữ liệu để xuất Excel.")
        return None

    required_columns = ['POS', 'DVUT', 'TENDV']
    if not all(col in df.columns for col in required_columns):
        st.error("Dữ liệu cần có các cột: POS, DVUT, TENDV")
        return None

    # Kiểm tra file mẫu tồn tại
    if not os.path.exists(template_file_path):
        st.error(f"File mẫu không tồn tại: {template_file_path}")
        return None
    
    try:
        wb = load_workbook(template_file_path)
        st.info(f"Đã đọc file mẫu: {template_file_path}")
    except Exception as e:
        st.error(f"Không thể đọc file mẫu: {e}")
        return None
    
    # Chuẩn hóa dữ liệu trước khi ghi
    df_cleaned = df.copy()
    string_columns = ['POS', 'DVUT', 'TENDV']
    for col in df_cleaned.columns:
        if col in string_columns:
            df_cleaned[col] = df_cleaned[col].astype(str).str.strip()
        else:
            if not df_cleaned[col].empty and df_cleaned[col].notna().any():
                df_cleaned[col] = pd.to_numeric(df_cleaned[col], errors='coerce')

    template_sheet = wb[wb.sheetnames[0]]
    original_template_name = wb.sheetnames[0]
    
    pos_list = df_cleaned['POS'].unique()
    
    for idx, pos in enumerate(pos_list):
        pos_data = df_cleaned[df_cleaned['POS'] == pos]
        sheet_name = str(pos)[:31]
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        if idx == 0:
            new_sheet = template_sheet
            new_sheet.title = sheet_name
        else:
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name
            for row in new_sheet['A15':f'V{new_sheet.max_row}']:
                for cell in row:
                    cell.value = None
        
        for row_idx, row in enumerate(pos_data.itertuples(), start=15):
            new_sheet[f'A{row_idx}'] = row.DVUT
            new_sheet[f'B{row_idx}'] = row.TENDV
            new_sheet[f'C{row_idx}'] = row.SH_CHOVAY if hasattr(row, 'SH_CHOVAY') else None
            new_sheet[f'D{row_idx}'] = row.CHOVAY if hasattr(row, 'CHOVAY') else None
            new_sheet[f'E{row_idx}'] = row.LK_CHOVAY if hasattr(row, 'LK_CHOVAY') else None
            new_sheet[f'F{row_idx}'] = row.THUNO if hasattr(row, 'THUNO') else None
            new_sheet[f'G{row_idx}'] = row.LK_THUNO if hasattr(row, 'LK_THUNO') else None
            new_sheet[f'H{row_idx}'] = row.SOTO_DUNO if hasattr(row, 'SOTO_DUNO') else None
            new_sheet[f'I{row_idx}'] = row.SH_DUNO if hasattr(row, 'SH_DUNO') else None
            new_sheet[f'J{row_idx}'] = row.DUNO if hasattr(row, 'DUNO') else None
            new_sheet[f'K{row_idx}'] = row.SH_QHAN if hasattr(row, 'SH_QHAN') else None
            new_sheet[f'L{row_idx}'] = row.QHAN if hasattr(row, 'QHAN') else None
            new_sheet[f'M{row_idx}'] = row.TL_NQH if hasattr(row, 'TL_NQH') else None
            new_sheet[f'N{row_idx}'] = row.TG_QHAN if hasattr(row, 'TG_QHAN') else None
            new_sheet[f'O{row_idx}'] = row.KHOANH if hasattr(row, 'KHOANH') else None
            new_sheet[f'P{row_idx}'] = row.LKSH_CHOVAY if hasattr(row, 'LKSH_CHOVAY') else None
            new_sheet[f'Q{row_idx}'] = row.TG_DUNO if hasattr(row, 'TG_DUNO') else None
            new_sheet[f'R{row_idx}'] = row.TYLE if hasattr(row, 'TYLE') else None
            new_sheet[f'S{row_idx}'] = row.TEN if hasattr(row, 'TEN') else None
            new_sheet[f'T{row_idx}'] = row.NGAYBC if hasattr(row, 'NGAYBC') else None
            new_sheet[f'U{row_idx}'] = row.TENPGD if hasattr(row, 'TENPGD') else None
            new_sheet[f'V{row_idx}'] = row.POS
    
    if len(pos_list) > 1 and original_template_name in wb.sheetnames:
        wb.remove(wb[original_template_name])
    
    # Tạo thư mục đầu ra nếu chưa tồn tại
    output_dir = os.path.dirname(output_file_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        st.info(f"Đã tạo thư mục: {output_dir}")
    
    try:
        wb.save(output_file_path)
        st.info(f"Đã lưu file: {output_file_path}")
        return output_file_path
    except Exception as e:
        st.error(f"Lỗi khi lưu file: {e}")
        return None