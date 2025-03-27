import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

def export_pos_to_excel(df, template_file_path, output_file_path="output_pos_data_fixed.xlsx"):
    # """
    # Xuất dữ liệu POS từ DataFrame ra file Excel với các sheet riêng cho từng POS.
    
    # Parameters:
    # - df (pd.DataFrame): DataFrame chứa dữ liệu với các cột 'POS', 'TENPGD', 'DUNO'.
    # - template_file_path (str): Đường dẫn đến file mẫu Excel.
    # - output_file_path (str): Đường dẫn đến file Excel kết quả.
    
    # Returns:
    # - str: Đường dẫn đến file kết quả nếu thành công, None nếu thất bại.
    # """
    required_columns = ['POS','CHTRINH','TENCT','TR_DUNO','DUNO','TRONGHAN','QUAHAN','KHOANH','DN_NGANHAN','DN_TRUNGHAN','DN_DAIHAN','CHOVAY','THUNO','GOCXOA']
    if not all(col in df.columns for col in required_columns):
        st.error("Dữ liệu cần có các cột:  POS,TENCT,TR_DUNO,DUNO,TRONGHAN,QUAHAN,KHOANH,DN_NGANHAN,DN_TRUNGHAN,DN_DAIHAN,CHOVAY,THUNO,GOCXOA")
        return None

    # Kiểm tra file mẫu tồn tại
    if not os.path.exists(template_file_path):
        st.error(f"File mẫu không tồn tại: {template_file_path}")
        return None
    
    try:
        wb = load_workbook(template_file_path)
        st.info(f"Đã đọc file mẫu: {template_file_path}")
    except Exception as e:
        st.error(f"Không thể đọc file mẫu: {e}")
        return None
    
    template_sheet = wb[wb.sheetnames[0]]
    original_template_name = wb.sheetnames[0]
    
    pos_list = df['POS'].unique()
    
    for idx, pos in enumerate(pos_list):
        pos_data = df[df['POS'] == pos]
        sheet_name = str(pos)[:31]
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        if idx == 0:
            new_sheet = template_sheet
            new_sheet.title = sheet_name
        else:
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name
            for row in new_sheet['A15':f'V{new_sheet.max_row}']:
                for cell in row:
                    cell.value = None
        
        for row_idx, row in enumerate(pos_data.itertuples(), start=15):
            new_sheet[f'A{row_idx}'] = row.CHTRINH
            new_sheet[f'B{row_idx}'] = row.TENCT
            new_sheet[f'C{row_idx}'] = row.CHOVAY
            new_sheet[f'D{row_idx}'] = row.LK_CHOVAY
            new_sheet[f'E{row_idx}'] = row.THUNO
            new_sheet[f'F{row_idx}'] = row.LK_THUNO
            new_sheet[f'G{row_idx}'] = row.GOCXOA
            new_sheet[f'H{row_idx}'] = row.LK_GOCXOA
            new_sheet[f'I{row_idx}'] = row.TR_DUNO
            new_sheet[f'J{row_idx}'] = row.DUNO
            new_sheet[f'K{row_idx}'] = row.TG_DUNO
            new_sheet[f'L{row_idx}'] = row.TRONGHAN
            new_sheet[f'M{row_idx}'] = row.QUAHAN
            new_sheet[f'N{row_idx}'] = row.KHOANH
            new_sheet[f'O{row_idx}'] = row.DN_NGANHAN
            new_sheet[f'P{row_idx}'] = row.DN_TRUNGHAN
            new_sheet[f'Q{row_idx}'] = row.DN_DAIHAN
            new_sheet[f'R{row_idx}'] = row.SH_DUNO
            new_sheet[f'S{row_idx}'] = row.LKSH_CHOVAY
            new_sheet[f'T{row_idx}'] = row.BQ_DUNO
            new_sheet[f'U{row_idx}'] = row.TEN
            new_sheet[f'V{row_idx}'] = row.NGAYBC
        
        #new_sheet.freeze_panes = 'B2'
    
    if len(pos_list) > 1 and original_template_name in wb.sheetnames:
        wb.remove(wb[original_template_name])
    
    # Tạo thư mục đầu ra nếu chưa tồn tại
    output_dir = os.path.dirname(output_file_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        st.info(f"Đã tạo thư mục: {output_dir}")
    
    try:
        wb.save(output_file_path)
        st.info(f"Đã lưu file: {output_file_path}")
        return output_file_path
    except Exception as e:
        st.error(f"Lỗi khi lưu file: {e}")
        return None

# Giao diện Streamlit
#st.title("Xuất dữ liệu POS ra Excel với cố định dòng cột")

# Định nghĩa DataFrame trực tiếp
data = {
    'POS': [2201, 2201, 2201, 2202, 2202, 2202],
    'TENPGD': ['Thành phố', 'Thành phố', 'Thành phố', 'Cao Lộc', 'Cao Lộc', 'Cao Lộc'],
    'DUNO': [238.0, 924.0, 170897.0, 37742.0, 1665.0, 123145.0]
}
df = pd.DataFrame(data)

# Đường dẫn file mẫu
TEMPLATE_FILE_PATH = "template.xlsx"  # Thay bằng đường dẫn thực tế

# Tùy chọn tên file đầu ra
# output_file_name = st.text_input("Tên file đầu ra", value="test_export_24_02_2025.xlsx")

# # Nút bắt đầu ghi dữ liệu
# if st.button("Bắt đầu ghi dữ liệu"):
#     with st.spinner("Đang xử lý..."):
#         output_file = export_pos_to_excel(df, TEMPLATE_FILE_PATH, output_file_name)
#         if output_file:
#             try:
#                 with open(output_file, "rb") as file:
#                     st.download_button(
#                         label="Tải file Excel kết quả",
#                         data=file,
#                         file_name=os.path.basename(output_file),
#                         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                     )
#                 st.success("Xuất dữ liệu thành công với dòng cột đã cố định!")
#             except Exception as e:
#                 st.error(f"Không thể mở file để tải: {e}")
#         else:
#             st.warning("Không tạo được file đầu ra, kiểm tra thông báo lỗi ở trên.")

        # st.markdown("""
        # ### Hướng dẫn:
        # 1. Đảm bảo file mẫu (`template.xlsx`) đã được đặt đúng đường dẫn.
        # 2. Nhập tên file đầu ra (nếu muốn thay đổi).
        # 3. Nhấn nút "Bắt đầu ghi dữ liệu" để xử lý.
        # 4. Tải file kết quả bằng nút "Tải file Excel kết quả" sau khi xử lý xong.
        # """)